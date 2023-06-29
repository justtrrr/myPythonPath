import tkinter as tk
from tkinter import filedialog
import re
from collections import Counter
from openpyxl import Workbook
from openpyxl.chart import BarChart, Reference
import os
import os.path


def open_file_dialog():
    root = tk.Tk()
    root.withdraw()
    return filedialog.askopenfilename(filetypes=[("Text files", "*.txt")])


def process_letters(file_path):
    with open(file_path, 'r', encoding='utf-8') as file:
        text = file.read().lower()

    letters = re.findall(r'[а-я]', text)

    letter_counts = Counter(letters)

    return sorted(letter_counts.items(), key=lambda x: x[0])


def process_words(file_path):
    with open(file_path, 'r', encoding='utf-8') as file:
        text = file.read().lower()

    text = re.sub(r'\d+', ' ', text)
    text = re.sub(r'[^\w\s]', ' ', text)

    word_counts = Counter(text.split())

    return sorted(word_counts.items(), key=lambda x: (-x[1], x[0]))


def write_to_excel(filename, sorted_letter_counts, sorted_words_counts):
    wb = Workbook()
    ws1 = wb.active
    ws1.title = "Буквы"
    ws1.cell(row=1, column=1, value="Буква")
    ws1.cell(row=1, column=2, value="Количество")
    ws1.cell(row=1, column=3, value="Вероятность")
    ws1.cell(row=1, column=4, value="Энтропия")
    ws1.cell(row=1, column=5, value="Энтропия в натах")

    for i, (letter, count) in enumerate(sorted_letter_counts):
        ws1.cell(row=i+2, column=1, value=letter)
        ws1.cell(row=i+2, column=2, value=count)
        ws1.cell(row=i+2, column=3, value=f"=B{i+2}/B{len(sorted_letter_counts)+2}")
        ws1.cell(row=i + 2, column=4, value=f"=-C{i+2}*LOG(C{i+2}, 2)")
        ws1.cell(row=i + 2, column=5, value=f"=-C{i + 2}*LOG(C{i + 2}, exp(1))")

    chart1 = BarChart()
    chart1.title = "Вероятность букв"
    chart1.x_axis.title = "Буква"
    chart1.y_axis.title = "Вероятность"
    data = Reference(ws1, min_col=3, min_row=2, max_row=len(sorted_letter_counts) + 1)
    categories = Reference(ws1, min_col=1, min_row=2, max_row=len(sorted_letter_counts) + 1)
    chart1.add_data(data)
    chart1.set_categories(categories)
    chart1.legend = None
    ws1.add_chart(chart1, "G1")

    chart2 = BarChart()
    chart2.title = "Энтропия букв"
    chart2.x_axis.title = "Буква"
    chart2.y_axis.title = "Энтропия"
    data2 = Reference(ws1, min_col=4, min_row=2, max_row=len(sorted_letter_counts) + 1)
    categories2 = Reference(ws1, min_col=1, min_row=2, max_row=len(sorted_letter_counts) + 1)
    chart2.add_data(data2)
    chart2.set_categories(categories2)
    chart2.legend = None
    ws1.add_chart(chart2, "G16")

    chart3 = BarChart()
    chart3.title = "Энтропия в натах"
    chart3.x_axis.title = "Буква"
    chart3.y_axis.title = "Энтропия в натах"
    data3 = Reference(ws1, min_col=5, min_row=2, max_row=len(sorted_letter_counts) + 1)
    categories3 = Reference(ws1, min_col=1, min_row=2, max_row=len(sorted_letter_counts) + 1)
    chart3.add_data(data3)
    chart3.set_categories(categories3)
    chart3.legend = None
    ws1.add_chart(chart3, "G31")

    last_row = len(sorted_letter_counts) + 2
    ws1.cell(row=last_row, column=1, value="Всего")
    ws1.cell(row=last_row, column=2, value=f"=SUM(B2:B{last_row-1})")

    ws2 = wb.create_sheet("Слова")
    wb.active = ws2
    ws2.cell(row=1, column=1, value="Слово")
    ws2.cell(row=1, column=2, value="Количество")
    ws2.cell(row=1, column=3, value="Всего")

    for i, (word, count) in enumerate(sorted_words_counts):
        ws2.cell(row=i+2, column=1, value=word)
        ws2.cell(row=i+2, column=2, value=count)

    last_row = len(sorted_words_counts) + 2
    ws2.cell(row=2, column=3, value=f"=SUM(B2:B{last_row-1})")

    wb.save(filename)


def program(filename):
    os.system(f'start "" "{filename}"')


def main():
    file_path = open_file_dialog()
    sorted_letter_counts = process_letters(file_path)
    sorted_words_counts = process_words(file_path)
    filename = f"Анализ {os.path.splitext(os.path.basename(file_path))[0]}.xlsx"
    write_to_excel(filename, sorted_letter_counts, sorted_words_counts)
    program(filename)


if __name__ == '__main__':
    main()
